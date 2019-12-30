## Import Statements
from sys import argv, exit
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import datetime
from openpyxl.styles.borders import Border, Side
from openpyxl import utils
from openpyxl import styles
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
#datetime_style = styles.Style(number_format="YYYY-MM-DD HH:MM:SS")
#duration_style = styles.Style(number_format="HH:MM:SS")
#date_style = styles.Style(number_format="YYYY-MM-DD")
datetime_style = styles.named_styles.NamedStyle(name = 'datetime_style', number_format="YYYY-MM-DD HH:MM:SS")
duration_style = styles.named_styles.NamedStyle(name = 'duration_style', number_format="HH:MM:SS")
date_style = styles.named_styles.NamedStyle(name = 'date_style', number_format="YYYY-MM-DD")

## Constants and Names
# filenamedate = "170525" #file got corrupted
filenamedate = "191227"
workbookname_type = "_Time_Management_"
fileextension = ".xlsx"
logworkbookName = [filenamedate, workbookname_type, "Log", fileextension]
goalworkbookName = [filenamedate, workbookname_type, "Goal", fileextension]
logworkbookName = str(''.join(logworkbookName))
goalworkbookName = str(''.join(goalworkbookName))
timeSheet = "Time"
goalSheet = "Goal"
summarySheet = "Summary"
timeSubject = "A"
timeTextbook = "B"
timeActivity = "C"
timeChapter = "D"
timeStart = "E"
timeEnd = "F"
timeDuration = "G"
goalSubject = "A"
goalTextbook = "B"
goalActivity = "C"
goalChapter = "D"
goalDuration = "E"
exerciseDate = "A"
exerciseType = "B"
exerciseCalories = "C"
exerciseDuration = "D"
summaryDate = "A"
summaryDuration = "B"
summaryPercentage = "C"

def start(workbookName_input=logworkbookName, timeSheet_input=timeSheet, subject="Subject", textbook="", activity="", chapter=""):
    """Open workbook defined in the update_time.py script, and update the Time sheet,
     with the latest row stating the type of work you put in as parameters and the current time as start time"""
    wb = load_workbook(workbookName_input)
    try:
        wb.add_named_style(datetime_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(duration_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(date_style)
    except ValueError:
        pass
    print("\n\tWorkbook", workbookName_input)
    ws = wb[timeSheet_input]
    print("\thas sheets, ", wb.sheetnames)
    print("\tOf which you are editing ", ws)
    lastRow = ws.max_row
    lastEndCell = timeEnd + str(lastRow)
    if(lastRow == 1):
        lastEndCell = timeEnd + str(lastRow+1)
    if ws[lastEndCell].value == None and lastRow > 1:
        print("\nYou have currently an ongoing activity. Check the excel file, input correct parameters, and end your current activity before starting a new one.")
        print("\n\tws[lastEndCell].value = ", ws[lastEndCell].value)
        exit()     
    subjectCell = timeSubject + str(lastRow+1)
    textbookCell = timeTextbook + str(lastRow+1)
    activityCell = timeActivity + str(lastRow+1)
    chapterCell = timeChapter + str(lastRow+1)
    startCell = timeStart + str(lastRow+1)
    currentTime = datetime.datetime.now()
    ws[subjectCell].value = subject
    ws[textbookCell].value = textbook
    ws[activityCell].value = activity
    ws[chapterCell].value = chapter
    #startTime = datetime.datetime.strptime(str(currentTime), '%Y-%m-%d %H:%M:%S.%f')
    ws[startCell] = datetime.datetime.strftime(currentTime, '%Y-%m-%d %H:%M:%S')
    ws[startCell].style = 'datetime_style'
    ws[subjectCell].border = thin_border
    ws[textbookCell].border = thin_border
    ws[activityCell].border = thin_border
    ws[chapterCell].border = thin_border
    ws[startCell].border = thin_border
    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
    wb.save(workbookName_input)

def end(workbookName_input=logworkbookName, timeSheet_input=timeSheet, subject="Subject", textbook="", activity="", chapter=""):
    """Open workbook defined in the update_time.py script, and update the Time sheet,
     with the latest row stating the type of work you put in as parameters and the current time as finish time"""
    wb = load_workbook(workbookName_input)
    try:
        wb.add_named_style(datetime_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(duration_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(date_style)
    except ValueError:
        pass
    print("\n\tWorkbook", workbookName_input)
    ws = wb[timeSheet_input]
    print("\thas sheets, ", wb.sheetnames)
    print("\tOf which you are editing ", ws)
    lastRow = ws.max_row
    subjectCell = timeSubject + str(lastRow)
    textbookCell = timeTextbook + str(lastRow)
    activityCell = timeActivity + str(lastRow)
    chapterCell = timeChapter + str(lastRow)
    startCell = timeStart + str(lastRow)
    endCell = timeEnd + str(lastRow)
    durationCell = timeDuration + str(lastRow)
    currentTime = datetime.datetime.now()
    subjectCompare = False
    textbookCompare = False
    activityCompare = False
    chapterCompare = False
    if (ws[subjectCell].value == subject) or (ws[subjectCell].value == str(subject)) or (ws[subjectCell].value == int(subject)) or (ws[subjectCell].value == float(subject)):
        subjectCompare = True
    if (ws[textbookCell].value == textbook) or (ws[textbookCell].value == str(textbook)) or (ws[textbookCell].value == int(textbook)) or (ws[textbookCell].value == float(textbook)):
        textbookCompare = True
    if (ws[activityCell].value == activity) or (ws[activityCell].value == str(activity)) or (ws[activityCell].value == int(activity)) or (ws[activityCell].value == float(activity)):
        activityCompare = True
    if (ws[chapterCell].value == chapter) or (ws[chapterCell].value == str(chapter)) or (ws[chapterCell].value == int(chapter)) or (ws[chapterCell].value == float(chapter)):
        chapterCompare = True
    #if ((ws[subjectCell].value == subject and ws[textbookCell].value == textbook) and (ws[activityCell].value == activity and ws[chapterCell].value == chapter)):
    #    print("\nYour input parameters do not match the current ongoing activity. Check the excel file and input correct parameters.")
    #    print("\n\tCurrent Values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    #    print("\n\tYour input values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", subject)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", textbook)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", activity)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", chapter)
    #    print("\n\tProceding")
    #elif ((ws[subjectCell].value == subject and ws[textbookCell].value == textbook) and (float(ws[activityCell].value) == activity and ws[chapterCell].value == chapter)):
    #    print("\nYour input parameters do not match the current ongoing activity. Check the excel file and input correct parameters.")
    #    print("\n\tCurrent Values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    #    print("\n\tYour input values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", subject)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", textbook)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", activity)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", chapter)
    #    print("\n\tProceding")
    #elif ((ws[subjectCell].value == subject and ws[textbookCell].value == textbook) and (ws[activityCell].value == activity and float(ws[chapterCell].value) == chapter)):
    #    print("\nYour input parameters do not match the current ongoing activity. Check the excel file and input correct parameters.")
    #    print("\n\tCurrent Values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    #    print("\n\tYour input values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", subject)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", textbook)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", activity)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", chapter)
    #    print("\n\tProceding")
    #elif ((ws[subjectCell].value == subject and ws[textbookCell].value == textbook) and (float(ws[activityCell].value) == activity and float(ws[chapterCell].value) == chapter)):
    #    print("\nYour input parameters do not match the current ongoing activity. Check the excel file and input correct parameters.")
    #    print("\n\tCurrent Values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    #    print("\n\tYour input values are: ")
    #    print("\n\t\tcell[", subjectCell, "] (subject) = ", subject)
    #    print("\n\t\tcell[", textbookCell, "] (textbook) = ", textbook)
    #    print("\n\t\tcell[", activityCell, "] (activity) = ", activity)
    #    print("\n\t\tcell[", chapterCell, "] (chapter) = ", chapter)
    #    print("\n\tProceding")
    if ( (subjectCompare == True) and (subjectCompare == True) and (subjectCompare == True) and (subjectCompare == True) ):
        print("\nYour input parameters match the current ongoing activity.")
        print("\n\tCurrent Values are: ")
        print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value, " of type ", type(ws[subjectCell].value))
        print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value,  " of type ", type(ws[textbookCell].value))
        print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value,  " of type ", type(ws[activityCell].value))
        print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value,  " of type ", type(ws[chapterCell].value))
        print("\n\tYour input values are: ")
        print("\n\t\tcell[", subjectCell, "] (subject) = ", subject,  " of type ", type(subject))
        print("\n\t\tcell[", textbookCell, "] (textbook) = ", textbook,  " of type ", type(textbook))
        print("\n\t\tcell[", activityCell, "] (activity) = ", activity,  " of type ", type(activity))
        print("\n\t\tcell[", chapterCell, "] (chapter) = ", chapter,  " of type ", type(chapter))
        print("\n\tProceeding")   
    else:
        print("\nYour input parameters do not match the current ongoing activity. Check the excel file and input correct parameters.")
        print("\n\tCurrent Values are: ")
        print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value, " of type ", type(ws[subjectCell].value))
        print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value,  " of type ", type(ws[textbookCell].value))
        print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value,  " of type ", type(ws[activityCell].value))
        print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value,  " of type ", type(ws[chapterCell].value))
        print("\n\tYour input values are: ")
        print("\n\t\tcell[", subjectCell, "] (subject) = ", subject,  " of type ", type(subject))
        print("\n\t\tcell[", textbookCell, "] (textbook) = ", textbook,  " of type ", type(textbook))
        print("\n\t\tcell[", activityCell, "] (activity) = ", activity,  " of type ", type(activity))
        print("\n\t\tcell[", chapterCell, "] (chapter) = ", chapter,  " of type ", type(chapter))
        print("\n\tEXITING")
        exit()
    #ws[endCell].value = utils.datetime.to_excel(currentTime)
    ws[endCell] = datetime.datetime.strftime(currentTime, '%Y-%m-%d %H:%M:%S')
    #ws[subjectCell].border = thin_border
    #ws[textbookCell].border = thin_border
    #ws[activityCell].border = thin_border
    #ws[chapterCell].border = thin_border
    #ws[startCell].border = thin_border
    ws[endCell].border = thin_border
    #startTime = utils.datetime.from_excel(ws[startCell].value)
    #startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
    try:
        startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
    except ValueError:
        startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S.%f')
    print("\nstartTime = ", startTime)
    print("\nendTime = ", currentTime)
    #durationTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S') - datetime.datetime.strptime(str(startTime), '%Y-%m-%d %H:%M:%S')
    durationTime = currentTime - startTime
    print("\ndurationTime = ", durationTime)
    #durationTime = datetime.datetime.strptime(str(durationTime), '%Y-%m-%d %H:%M:%S.%f')
    durationTime = datetime.datetime.strptime(str(ws["H1"].value), '%Y-%m-%d %H:%M:%S') + durationTime
    ws[durationCell].value = datetime.datetime.strftime(durationTime, '%H:%M:%S')
    ws[startCell].style = 'datetime_style'
    ws[endCell].style = 'datetime_style'
    ws[durationCell].style = 'duration_style'
    ws[startCell].border = thin_border
    ws[endCell].border = thin_border
    ws[durationCell].border = thin_border
    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
    print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
    print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
    wb.save(workbookName_input)

def show(workbookName_input=logworkbookName, timeSheet_input=timeSheet, row=0):
    """"""
    wb = load_workbook(workbookName_input)
    try:
        wb.add_named_style(datetime_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(duration_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(date_style)
    except ValueError:
        pass
    print("\n\tWorkbook", workbookName_input)
    ws = wb[timeSheet_input]
    print("\thas sheets, ", wb.sheetnames)
    print("\tOf which you are editing ", ws)
    if row==0:
        lastRow = ws.max_row
        for i in range(lastRow-1):
            show_time_line(workbookName_input=timeworkbookName, timeSheet_input=timeSheet, row=i)
            #subjectCell = timeSubject + str(i+2)
            #textbookCell = timeTextbook + str(i+2)
            #activityCell = timeActivity + str(i+2)
            #chapterCell = timeChapter + str(i+2)
            #startCell = timeStart + str(i+2)
            #endCell = timeEnd + str(i+2)
            #durationCell = timeDuration + str(i+2)
            #print("\n")
            #print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
            #print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
            #print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
            #print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
            #print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
            #print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
            #print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
    elif row==-1:
        lastRow = ws.max_row;
        show_time_line(workbookName_input=logworkbookName, timeSheet_input=timeSheet, row=(lastRow) )
        #subjectCell = timeSubject + str(lastRow)
        #textbookCell = timeTextbook + str(lastRow)
        #activityCell = timeActivity + str(lastRow)
        #chapterCell = timeChapter + str(lastRow)
        #startCell = timeStart + str(lastRow)
        #endCell = timeEnd + str(lastRow)
        #durationCell = timeDuration + str(lastRow)
        #print("\n")
        #print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
        #print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
        #print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
        #print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
        #print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
        #print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
        #print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
    elif row < -1:
        lastRow = ws.max_row;
        if lastRow - row + 1 < 1:
            print("\nYou row value of ", row, " corresponds to negative row value")
            exit()
        row_value = lastRow + row + 1;
        print("\nYour input = ", row, "\t| last row = ", lastRow, "\t| corresponding row = ", row_value)
        show_time_line(workbookName_input=logworkbookName, timeSheet_input=timeSheet, row=row_value)
        #subjectCell = timeSubject + str(row_value)
        #textbookCell = timeTextbook + str(row_value)
        #activityCell = timeActivity + str(row_value)
        #chapterCell = timeChapter + str(row_value)
        #startCell = timeStart + str(row_value)
        #endCell = timeEnd + str(row_value)
        #durationCell = timeDuration + str(row_value)
        #print("\n")
        #print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
        #print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
        #print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
        #print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
        #print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
        #print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
        #print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)

    elif row>1:
        show_time_line(workbookName_input=timeworkbookName, timeSheet_input=timeSheet, row=row)
        #subjectCell = timeSubject + str(row)
        #textbookCell = timeTextbook + str(row)
        #activityCell = timeActivity + str(row)
        #chapterCell = timeChapter + str(row)
        #startCell = timeStart + str(row)
        #endCell = timeEnd + str(row)
        #durationCell = timeDuration + str(row)
        #print("\n")
        #print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
        #print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
        #print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
        #print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
        #print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
        #print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
        #print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
    else:
        print("\nYou need row number == 0 for calculating all rows or row > 1 for calculating specific row.")
        exit()

def show_time_line(workbookName_input=logworkbookName, timeSheet_input=timeSheet, row=0):
    wb = load_workbook(workbookName_input)
    ws = wb[timeSheet_input]
    subjectCell = timeSubject + str(row)
    textbookCell = timeTextbook + str(row)
    activityCell = timeActivity + str(row)
    chapterCell = timeChapter + str(row)
    startCell = timeStart + str(row)
    endCell = timeEnd + str(row)
    durationCell = timeDuration + str(row)
    print("\n")
    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
    print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
    print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)

def duration(workbookName_input=logworkbookName, timeSheet_input=timeSheet, row=0):
    """"""
    wb = load_workbook(workbookName_input)
    try:
        wb.add_named_style(datetime_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(duration_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(date_style)
    except ValueError:
        pass
    print("\n\tWorkbook", workbookName_input)
    ws = wb[timeSheet_input]
    print("\thas sheets, ", wb.sheetnames)
    print("\tOf which you are editing ", ws)
    if row==0:
        lastRow = ws.max_row
        for i in range(lastRow-1):
            startCell = timeStart + str(i+2)
            endCell = timeEnd + str(i+2)
            durationCell = timeDuration + str(i+2)
            #startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
            #endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S')
            try:
                startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S.%f')
            try:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S.%f')
            durationTime = endTime - startTime
            #durationTime = datetime.datetime.strptime(str(endTime), '%Y-%m-%d %H:%M:%S') - datetime.datetime.strptime(str(startTime), '%Y-%m-%d %H:%M:%S')
            #durationTime = datetime.datetime.strptime(str(ws[durationCell].value), '%Y-%m-%d %H:%M:%S.%f')
            #durationTime = datetime.datetime.strftime(endTime, '%Y-%m-%d %H:%M:%S') - datetime.datetime.strftime(startTime, '%Y-%m-%d %H:%M:%S')
            durationTime = endTime - startTime
            durationTime = datetime.datetime.strptime(str(ws["H1"].value), '%Y-%m-%d %H:%M:%S') + durationTime
            ws[durationCell].value = datetime.datetime.strftime(durationTime, '%H:%M:%S')
            ws[endCell].style = 'datetime_style'
            ws[durationCell].style = 'duration_style'
            ws[startCell].border = thin_border
            ws[endCell].border = thin_border
            ws[durationCell].border = thin_border
            print("\n")
            print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
            print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
            print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
            wb.save(workbookName_input)
    elif row==-1:
        lastRow = ws.max_row
        subjectCell = timeSubject + str(lastRow)
        textbookCell = timeTextbook + str(lastRow)
        activityCell = timeActivity + str(lastRow)
        chapterCell = timeChapter + str(lastRow)
        startCell = timeStart + str(lastRow)
        endCell = timeEnd + str(lastRow)
        durationCell = timeDuration + str(lastRow)
        try:
            startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
        except ValueError:
            startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S.%f')
        if ws[endCell].value == None:
            endTime = datetime.datetime.now()
        else:
            try:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S.%f')
        durationTime = endTime - startTime
        #durationTime = datetime.datetime.strptime(str(endTime), '%Y-%m-%d %H:%M:%S.%f') - datetime.datetime.strptime(str(startTime), '%Y-%m-%d %H:%M:%S.%f')
        #durationTime = datetime.datetime.strptime(str(ws[durationCell].value), '%Y-%m-%d %H:%M:%S.%f')
        durationTime = datetime.datetime.strptime(str(ws["H1"].value), '%Y-%m-%d %H:%M:%S') + durationTime
        ws[durationCell].value = datetime.datetime.strftime(durationTime, '%H:%M:%S')
        ws[startCell].style = 'datetime_style'
        ws[endCell].style = 'datetime_style'
        ws[durationCell].style = 'duration_style'
        ws[startCell].border = thin_border
        ws[endCell].border = thin_border
        ws[durationCell].border = thin_border
        print("\n")
        print("\n\t\tcell[", subjectCell, "] (start) = ", ws[subjectCell].value)
        print("\n\t\tcell[", textbookCell, "] (start) = ", ws[textbookCell].value)
        print("\n\t\tcell[", activityCell, "] (start) = ", ws[activityCell].value)
        print("\n\t\tcell[", chapterCell, "] (start) = ", ws[chapterCell].value)
        print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
        print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
        print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
        if ws[endCell].value != None:
            wb.save(workbookName_input)
    elif row < -1:
        lastRow = ws.max_row
        if lastRow - row + 1 < 1:
            print("\nYou row value of ", row, " corresponds to negative row value")
            exit()
        row_value = lastRow + row + 1
        print("\nYour input = ", row, "\t| last row = ", lastRow, "\t| corresponding row = ", row_value)
        subjectCell = timeSubject + str(row_value)
        textbookCell = timeTextbook + str(row_value)
        activityCell = timeActivity + str(row_value)
        chapterCell = timeChapter + str(row_value)
        startCell = timeStart + str(row_value)
        endCell = timeEnd + str(row_value)
        durationCell = timeDuration + str(row_value)
        try:
            startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
        except ValueError:
            startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S.%f')
        if ws[endCell].value == None:
            endTime = datetime.datetime.now()
        else:
            try:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S.%f')
        durationTime = endTime - startTime
        #durationTime = datetime.datetime.strptime(str(endTime), '%Y-%m-%d %H:%M:%S.%f') - datetime.datetime.strptime(str(startTime), '%Y-%m-%d %H:%M:%S.%f')
        #durationTime = datetime.datetime.strptime(str(ws[durationCell].value), '%Y-%m-%d %H:%M:%S.%f')
        durationTime = datetime.datetime.strptime(str(ws["H1"].value), '%Y-%m-%d %H:%M:%S') + durationTime
        ws[durationCell].value = datetime.datetime.strftime(durationTime, '%H:%M:%S')
        ws[startCell].style = 'datetime_style'
        ws[endCell].style = 'datetime_style'
        ws[durationCell].style = 'duration_style'
        ws[startCell].border = thin_border
        ws[endCell].border = thin_border
        ws[durationCell].border = thin_border
        print("\n")
        print("\n\t\tcell[", subjectCell, "] (start) = ", ws[subjectCell].value)
        print("\n\t\tcell[", textbookCell, "] (start) = ", ws[textbookCell].value)
        print("\n\t\tcell[", activityCell, "] (start) = ", ws[activityCell].value)
        print("\n\t\tcell[", chapterCell, "] (start) = ", ws[chapterCell].value)
        print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
        print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
        print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
        if ws[endCell].value != None:
            wb.save(workbookName_input)
    elif row>1:
        startCell = timeStart + str(row)
        endCell = timeEnd + str(row)
        durationCell = timeDuration + str(row)
        try:
            startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S')
        except ValueError:
            startTime = datetime.datetime.strptime(str(ws[startCell].value), '%Y-%m-%d %H:%M:%S.%f')
        if ws[endCell].value == None:
            endTime = datetime.datetime.now()
        else:
            try:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                endTime = datetime.datetime.strptime(str(ws[endCell].value), '%Y-%m-%d %H:%M:%S.%f')
        durationTime = endTime - startTime
        #durationTime = datetime.datetime.strptime(str(endTime), '%Y-%m-%d %H:%M:%S.%f') - datetime.datetime.strptime(str(startTime), '%Y-%m-%d %H:%M:%S.%f')
        #durationTime = datetime.datetime.strptime(str(ws[durationCell].value), '%Y-%m-%d %H:%M:%S.%f')
        durationTime = datetime.datetime.strptime(str(ws["H1"].value), '%Y-%m-%d %H:%M:%S') + durationTime
        ws[durationCell].value = datetime.datetime.strftime(durationTime, '%H:%M:%S')
        ws[startCell].style = 'datetime_style'
        ws[endCell].style = 'datetime_style'
        ws[durationCell].style = 'duration_style'
        ws[startCell].border = thin_border
        ws[endCell].border = thin_border
        ws[durationCell].border = thin_border
        print("\n\t\tcell[", startCell, "] (start) = ", ws[startCell].value)
        print("\n\t\tcell[", endCell, "] (end) = ", ws[endCell].value)
        print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)
        if ws[endCell].value != None:
            wb.save(workbookName_input)
    else:
        print("\nYou need row number == 0 for calculating all rows or row > 1 for calculating specific row.")
        exit()

#def exercise(type, workbookName_input=workbookName, exerciseSheet_input=exerciseSheet, calories=0, duration=0):
#    """"""
#    """Open workbook defined in the update_time.py script, and update the Time sheet,
#     with the latest row stating the type of work you put in as parameters and the current time as start time"""
#    wb = load_workbook(workbookName_input)
#    print("\n\tWorkbook", workbookName_input)
#    ws = wb[exerciseSheet_input]
#    print("\thas sheets, ", wb.sheetnames)
#    print("\tOf which you are editing ", ws)
#    lastRow = ws.max_row
#    dateCell = exerciseDate + str(lastRow+1)
#    typeCell = exerciseType + str(lastRow+1)
#    caloriesCell = exerciseCalories + str(lastRow+1)
#    durationCell = exerciseDuration + str(lastRow+1)
#    currentTime = datetime.datetime.now()
#    ws[dateCell].value = utils.datetime.to_excel(currentTime)
#    ws[typeCell].value = type
#    ws[caloriesCell].value = calories
#    ws[durationCell].value = duration
#    ws[dateCell].style = date_style
#    ws[durationCell].style = duration_style
#    ws[dateCell].border = thin_border
#    ws[typeCell].border = thin_border
#    ws[caloriesCell].border = thin_border
#    ws[durationCell].border = thin_border

def start_goal(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, subject="Subject", textbook="", activity="", chapter=""):
    wb = load_workbook(workbookName_input)
    try:
        wb.add_named_style(datetime_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(duration_style)
    except ValueError:
        pass
    try:
        wb.add_named_style(date_style)
    except ValueError:
        pass
    print("\n\tWorkbook", workbookName_input)
    ws = wb[goalSheet_input]
    print("\thas sheets, ", wb.sheetnames)
    print("\tOf which you are editing ", ws)
    lastRow = ws.max_row
    for i in range(lastRow):
        subjectCell = goalSubject + str(i+1)
        textbookCell = goalTextbook + str(i+1)
        activityCell = goalActivity + str(i+1)
        chapterCell = goalChapter + str(i+1)
        if ws[subjectCell].value == subject:
            if ws[textbookCell].value == textbook:
                if ws[activityCell].value == activity:
                    if ws[chapterCell].value == chapter:
                        print("You already have a row with exactly the same input as your inputs")
                    else:
                        break
                else:
                    break
            else:
                break
        else:
            break
    subjectCell = goalSubject + str(lastRow+1)
    textbookCell = goalTextbook + str(lastRow+1)
    activityCell = goalActivity + str(lastRow+1)
    chapterCell = goalChapter + str(lastRow+1)
    ws[subjectCell].value = subject
    ws[textbookCell].value = textbook
    ws[activityCell].value = activity
    ws[chapterCell].value = chapter
    ws[subjectCell].border = thin_border
    ws[textbookCell].border = thin_border
    ws[activityCell].border = thin_border
    ws[chapterCell].border = thin_border
    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    wb.save(workbookName_input)

def duration_goal(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, workbookName_data=logworkbookName, goalSheet_data=timeSheet, subject="Subject", textbook="", activity="", chapter=""):
    """"""
    wb = load_workbook(workbookName_input)
    wb_data = load_workbook(workbookName_data)
    print("\n\tWorkbook ", workbookName_input)
    print("\n\tWorkbook to get data from ", workbookName_data)
    ws = wb[goalSheet_input]
    ws_data = wb_data[goalSheet_data]
    print("\thas sheets, ", wb.sheetnames)
    print("\tOf which you are editing ", ws)
    lastRow = ws.max_row
    lastRow_data = ws_data.max_row
    for i in range(lastRow+1):
        if i <= 1:
            continue
        time_sum = datetime.datetime.strptime("1900-01-01 00:00:00", '%Y-%m-%d %H:%M:%S')
        time_sum = datetime.datetime.time(time_sum)
        #time_sum = datetime.datetime.strptime(str(ws["F1"].value), '%Y-%m-%d %H:%M:%S')
        subjectCell = goalSubject + str(i)
        textbookCell = goalTextbook + str(i)
        activityCell = goalActivity + str(i)
        chapterCell = goalChapter + str(i)
        durationCell = goalDuration + str(i)
        for j in range(lastRow_data+1):
            if j <= 1:
                continue
            subjectCell_data = timeSubject + str(j)
            textbookCell_data = timeTextbook + str(j)
            activityCell_data = timeActivity + str(j)
            chapterCell_data = timeChapter + str(j)
            durationCell_data = timeDuration + str(j)
            print("i = ", i, "\tj = ", j, "\tSubject data = ", ws_data[subjectCell_data].value, "\tSubject Goal = ", ws[subjectCell].value)
            if ws_data[subjectCell_data].value == ws[subjectCell].value:
                #print("Subject Match: ", ws_data[subjectCell_data].value)
                if ws_data[textbookCell_data].value == ws[textbookCell].value:
                    #print("Textbook Match: ", ws_data[textbookCell_data].value)
                    if ws_data[activityCell_data].value == ws[activityCell].value:
                        #print("Acitivity Match: ", ws_data[activityCell_data].value)
                        if ws_data[chapterCell_data].value == ws[chapterCell].value:
                            #print("Chapter Match: ", ws_data[chapterCell_data].value)
                            print("match, timesum = ", time_sum)
                            time_sum = datetime.datetime.time(datetime.datetime.strptime("1900-01-01 "+str(ws_data[durationCell_data].value), '%Y-%m-%d %H:%M:%S')) + time_sum
                            #time_sum = time_sum + datetime.datetime.strptime(datetime.datetime.strptime(str(ws_data[durationCell_data].value), '%H:%M:%S'), '%Y-%m-%d %H:%M:%S')
                        else:
                            continue
                    else:
                        continue
                else:
                    continue
            else:
                continue
            ws[durationCell].value = datetime.datetime.strftime(time_sum, '%H:%M:%S')
        #print("line i = ", i)
        #show_goal_line(workbookName_input=workbookName_input, goalSheet_input=goalSheet_input, row=i)
    wb.save(workbookName_input)

def show_goal_line(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, row=0):
    wb = load_workbook(workbookName_input)
    ws = wb[goalSheet_input]
    subjectCell = goalSubject + str(row)
    textbookCell = goalTextbook + str(row)
    activityCell = goalActivity + str(row)
    chapterCell = goalChapter + str(row)
    durationCell = goalDuration + str(row)
    print("\n")
    print("\n\t\tcell[", subjectCell, "] (subject) = ", ws[subjectCell].value)
    print("\n\t\tcell[", textbookCell, "] (textbook) = ", ws[textbookCell].value)
    print("\n\t\tcell[", activityCell, "] (activity) = ", ws[activityCell].value)
    print("\n\t\tcell[", chapterCell, "] (chapter) = ", ws[chapterCell].value)
    print("\n\t\tcell[", durationCell, "] (duration) = ", ws[durationCell].value)

def show_goal(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, row=0):
    """"""
    wb = load_workbook(workbookName_input)
    ws = wb[goalSheet_input]
    lastRow = ws.max_row
    row = int(row)
    print("\nlast_row = ",lastRow)
    if row == 0:
        for i in range(lastRow+1):
            if i > 1:
                show_goal_line(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, row=i)
    if row < 0:
            show_goal_line(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, row=lastRow)
    if row > 1:
            show_goal_line(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, row=row)


print("\nargument inputs are: ", argv)
if argv[1] == "start" or argv[1] == "end":
    input = []
    for i in range(len(argv)-2):
        input.append(argv[i+2])
    for j in range(4-len(input)):
        input.append("")
    print("\nYour input file that goes into start or end is: ", input)
    if argv[1] == "start":
        start(subject=input[0], textbook=input[1], activity=input[2], chapter=input[3])
    if argv[1] == "end":
        end(subject=input[0], textbook=input[1], activity=input[2], chapter=input[3])
elif argv[1] == "duration":
    if len(argv)==3:
        duration(row=int(argv[2]))
    else:
        duration()
elif argv[1] == "show":
    if len(argv)==3:
        show(row=int(argv[2]))
    else:
        show()
elif argv[1] == "start_goal":
    input = []
    for i in range(len(argv)-2):
        input.append(argv[i+2])
    for j in range(4-len(input)):
        input.append("")
    print("\nYour input file that goes into start or end is: ", input)
    start_goal(subject=input[0], textbook=input[1], activity=input[2], chapter=input[3])
elif argv[1] == "duration_goal":
    #duration_goal()
    print("doing nothing because duration_goal has not been debuged yet")
elif (argv[1] == "show_goal" and len(argv) > 2):
    show_goal(workbookName_input=goalworkbookName, goalSheet_input=goalSheet, row=argv[2])